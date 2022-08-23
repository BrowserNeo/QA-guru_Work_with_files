import os
import shutil
import zipfile
import csv
from fileinput import filename

from PyPDF2 import PdfReader
from openpyxl import load_workbook


# Создаем папку и архив
if not os.path.exists('resources/tmp'):
    os.mkdir('resources/tmp')
    myzip = zipfile.ZipFile('resources/tmp/my_zip.zip', 'w')


# Добавляем файлы в архив
with zipfile.ZipFile('resources/tmp/my_zip.zip', 'w') as myzip:
    myzip.write('resources/username.csv')
    myzip.write('resources/file_example_XLSX_50.xlsx')
    myzip.write('resources/docs-pytest-org-en-latest.pdf')
    myzip.close()

# Проверка файла
def test_read_zip():
    for file_info in myzip.infolist():
        print(file_info.filename, file_info.date_time, file_info.file_size)

# Разархивируем файл
with zipfile.ZipFile('resources/tmp/my_zip.zip', 'r') as myzip:
    myzip.extractall('resources/tmp/')
    myzip.close()

# Проверка pdf
def test_read_pdf():
    reader = PdfReader('resources/tmp/resources/docs-pytest-org-en-latest.pdf')
    page = reader.pages[0]
    text = page.extract_text()
    assert 'pytest Documentation' in text

# Проверка xlsx
def test_read_xlsx():
    workbook = load_workbook('resources/tmp/resources/file_example_XLSX_50.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=2, column=1).value
    assert 'Dulce' == name

# Проверка csv
def test_read_csv_():
    with open('resources/tmp/resources/username.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'First_name' in str(headers)

# Удаляем папку


shutil.rmtree('resources/tmp')